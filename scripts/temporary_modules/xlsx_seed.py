#!/usr/bin/env python3
"""Fast XLSX helpers for temporary module seed imports.

Laravel keeps validation, permissions and database writes. This script only
does the expensive XLSX streaming work and returns compact JSON/CSV output.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import math
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


def stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "VERDADERO" if value else "FALSO"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer() and abs(value) < 1e15:
            return str(int(value))
        return f"{value:.6f}".rstrip("0").rstrip(".")
    if isinstance(value, dt.datetime):
        if value.time() == dt.time(0, 0, 0):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M:%S")

    return str(value).strip()


def open_workbook(path: str):
    return load_workbook(path, read_only=True, data_only=True)


def select_sheet(workbook, sheet_index: int):
    names = workbook.sheetnames
    if not names:
        raise ValueError("El archivo no contiene hojas.")
    index = max(0, min(sheet_index, len(names) - 1))
    return workbook[names[index]], names, index


def row_values(row) -> list[str]:
    return [stringify(cell.value) for cell in row]


def best_header_row(rows: dict[int, list[str]], max_candidate: int = 5) -> int:
    best_row = 1
    best_non_empty = 0
    for row_number in range(1, max_candidate + 1):
        values = rows.get(row_number, [])
        non_empty = sum(1 for value in values if value.strip() != "")
        if non_empty > best_non_empty:
            best_non_empty = non_empty
            best_row = row_number

    return best_row


def preview(args: argparse.Namespace) -> dict[str, Any]:
    workbook = open_workbook(args.file)
    try:
        sheet, sheet_names, sheet_index = select_sheet(workbook, args.sheet_index)
        header_row = max(1, args.header_row)

        scan_end = header_row + max(20, args.max_preview_rows)
        if header_row == 1:
            scan_end = max(scan_end, 25)

        rows: dict[int, list[str]] = {}
        for excel_row in sheet.iter_rows(min_row=1, max_row=scan_end):
            row_number = int(excel_row[0].row) if excel_row else 0
            if row_number < 1:
                continue
            rows[row_number] = row_values(excel_row)

        if header_row == 1:
            header_row = best_header_row(rows)

        labels = rows.get(header_row, [])
        data_rows = [cells for number, cells in rows.items() if number > header_row]
        max_cols = len(labels)
        for cells in data_rows:
            max_cols = max(max_cols, len(cells))
        max_cols = min(max_cols, args.max_columns)

        headers = []
        for index in range(max_cols):
            headers.append(
                {
                    "index": index,
                    "letter": get_column_letter(index + 1),
                    "label": labels[index] if index < len(labels) else "",
                }
            )

        preview_rows = []
        for number in sorted(rows):
            if number <= header_row:
                continue
            raw_cells = rows[number]
            cells = [(raw_cells[index] if index < len(raw_cells) else "") for index in range(max_cols)]
            if not any(cell.strip() for cell in cells):
                continue
            preview_rows.append({"row": number, "cells": cells})
            if len(preview_rows) >= args.max_preview_rows:
                break

        return {
            "success": True,
            "headers": headers,
            "suggested_map": {},
            "header_row": header_row,
            "sheet_names": sheet_names,
            "sheet_index": sheet_index,
            "preview_rows": preview_rows,
        }
    finally:
        workbook.close()


def export_csv(args: argparse.Namespace) -> dict[str, Any]:
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)

    workbook = open_workbook(args.file)
    written = 0
    try:
        sheet, _sheet_names, sheet_index = select_sheet(workbook, args.sheet_index)
        with output.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle, delimiter=",", quotechar='"', lineterminator="\n")
            for excel_row in sheet.iter_rows(min_row=max(1, args.data_start_row)):
                writer.writerow(row_values(excel_row))
                written += 1

        return {"success": True, "sheet_index": sheet_index, "rows": written, "output": str(output)}
    finally:
        workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="XLSX seed helper")
    subparsers = parser.add_subparsers(dest="command", required=True)

    preview_parser = subparsers.add_parser("preview")
    preview_parser.add_argument("--file", required=True)
    preview_parser.add_argument("--header-row", type=int, default=1)
    preview_parser.add_argument("--sheet-index", type=int, default=0)
    preview_parser.add_argument("--max-preview-rows", type=int, default=12)
    preview_parser.add_argument("--max-columns", type=int, default=100)

    export_parser = subparsers.add_parser("export-csv")
    export_parser.add_argument("--file", required=True)
    export_parser.add_argument("--sheet-index", type=int, default=0)
    export_parser.add_argument("--data-start-row", type=int, required=True)
    export_parser.add_argument("--output", required=True)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    try:
        if args.command == "preview":
            payload = preview(args)
        elif args.command == "export-csv":
            payload = export_csv(args)
        else:
            raise ValueError(f"Comando no soportado: {args.command}")
        print(json.dumps(payload, ensure_ascii=False))
        return 0
    except Exception as exc:
        print(json.dumps({"success": False, "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
