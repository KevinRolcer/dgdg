"""
Fotos en Excel dentro de la celda (ancla twoCell), no flotando encima.

Modos (consola):
  1 — Reanclar fotos ya pegadas en la columna indicada (recomendado si ya hay fotos).
  2 — Descargar desde URLs / =IMAGE() en esa columna e insertar.

También pide carpeta, nombre del .xlsx y columna (F, B, …).
"""
from __future__ import annotations

import os
import re
import sys
from io import BytesIO
from pathlib import Path

import requests
from openpyxl.drawing.image import Image
from openpyxl.utils import column_index_from_string
from PIL import Image as PILImage

from excel_imagenes_util import (
    anclar_imagen_dentro_celda,
    aplicar_lectura_tolerante_imagenes,
    contar_archivos_media,
    crear_respaldo_xlsx,
    guardar_workbook_con_respaldo,
    load_workbook_seguro,
    pedir_columna,
    reanclar_columna_en_xlsx,
)
from extract_img import pedir_ruta_excel

FILA_INICIO = 9
TAMANO_IMAGEN_PX = 100
TIMEOUT_DESCARGA = 30

_RE_IMAGE = re.compile(
    r'IMAGE\s*\(\s*["\']([^"\']+)["\']',
    re.IGNORECASE,
)
_RE_HYPERLINK = re.compile(
    r'HYPERLINK\s*\(\s*["\']([^"\']+)["\']',
    re.IGNORECASE,
)
_RE_URL = re.compile(r"https?://[^\s\"'<>]+", re.IGNORECASE)


def _pedir_modo() -> str:
    while True:
        print(
            "\nModo:\n"
            "  1 = Reanclar fotos existentes dentro de la celda (sin URLs)\n"
            "  2 = Insertar desde URLs o fórmulas en la columna\n"
        )
        m = input("Elige modo [1]: ").strip() or "1"
        if m in ("1", "2"):
            return m
        print("  Escribe 1 o 2.")


def _extraer_url(valor) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip() if not isinstance(valor, str) else valor.strip()
    if not texto:
        return None

    if texto.startswith("="):
        m = _RE_IMAGE.search(texto) or _RE_HYPERLINK.search(texto)
        if m:
            return m.group(1).strip()
        m = _RE_URL.search(texto)
        if m:
            return m.group(0).rstrip(")")
        return None

    if texto.lower().startswith(("http://", "https://")):
        return texto.split()[0]

    if os.path.isfile(texto):
        return texto

    return None


def _resolver_ruta_local(url_o_ruta: str, carpeta_excel: str) -> str | None:
    if url_o_ruta.lower().startswith(("http://", "https://")):
        return None
    if os.path.isabs(url_o_ruta) and os.path.isfile(url_o_ruta):
        return url_o_ruta
    candidato = os.path.join(carpeta_excel, url_o_ruta)
    if os.path.isfile(candidato):
        return candidato
    return None


def _descargar_o_abrir(url_o_ruta: str, carpeta_excel: str) -> bytes | None:
    local = _resolver_ruta_local(url_o_ruta, carpeta_excel)
    if local:
        try:
            return Path(local).read_bytes()
        except OSError:
            return None
    if not url_o_ruta.lower().startswith(("http://", "https://")):
        return None
    try:
        r = requests.get(url_o_ruta, timeout=TIMEOUT_DESCARGA)
        r.raise_for_status()
        return r.content
    except requests.RequestException:
        return None


def _a_jpeg_cuadrado(datos: bytes, lado: int) -> BytesIO | None:
    try:
        pil = PILImage.open(BytesIO(datos))
        if pil.mode not in ("RGB", "L"):
            pil = pil.convert("RGB")
        pil.thumbnail((lado, lado), PILImage.Resampling.LANCZOS)
        w, h = pil.size
        if w != h:
            cuad = PILImage.new("RGB", (lado, lado), (255, 255, 255))
            cuad.paste(pil, ((lado - w) // 2, (lado - h) // 2))
            pil = cuad
        buf = BytesIO()
        pil.save(buf, format="JPEG", quality=85)
        buf.seek(0)
        return buf
    except OSError:
        return None


def _quitar_imagenes_en_celda(ws, col_idx: int, row_idx: int) -> int:
    quitar = [
        img
        for img in ws._images
        if hasattr(img.anchor, "_from")
        and img.anchor._from.col == col_idx
        and img.anchor._from.row == row_idx
    ]
    for img in quitar:
        ws._images.remove(img)
    return len(quitar)


def _insertar_desde_urls(
    wb,
    col_letter: str,
    col_idx: int,
    carpeta_excel: str,
) -> dict[str, int]:
    from openpyxl.utils.units import pixels_to_EMU

    total = {"insertadas": 0, "sin_url": 0, "fallo_descarga": 0, "reemplazadas": 0}

    for ws in wb.worksheets:
        for fila in range(FILA_INICIO, ws.max_row + 1):
            url = _extraer_url(ws.cell(row=fila, column=col_idx).value)
            if not url:
                total["sin_url"] += 1
                continue

            datos = _descargar_o_abrir(url, carpeta_excel)
            if not datos:
                total["fallo_descarga"] += 1
                continue

            buf = _a_jpeg_cuadrado(datos, TAMANO_IMAGEN_PX)
            if buf is None:
                total["fallo_descarga"] += 1
                continue

            row_idx = fila - 1
            col0 = col_idx - 1
            total["reemplazadas"] += _quitar_imagenes_en_celda(ws, col0, row_idx)

            img_excel = Image(buf)
            anclar_imagen_dentro_celda(
                img_excel,
                col0,
                row_idx,
                pixels_to_EMU(TAMANO_IMAGEN_PX),
                pixels_to_EMU(TAMANO_IMAGEN_PX),
            )
            ws.row_dimensions[fila].height = max(
                ws.row_dimensions[fila].height or 0,
                TAMANO_IMAGEN_PX * 0.78,
            )
            ws.add_image(img_excel)
            total["insertadas"] += 1

    return total


def _confirmar_modo_urls_si_hay_imagenes(ruta: str) -> bool:
    n = contar_archivos_media(ruta)
    if n == 0:
        return True
    print(
        f"\n⚠ El libro ya tiene {n} archivo(s) en xl/media/.\n"
        "  El modo 2 reescribe el Excel con openpyxl y puede QUITAR imágenes\n"
        "  WMF o archivos dañados. Usa el modo 1 para solo reanclar.\n"
    )
    ok = input("¿Continuar con modo 2 de todos modos? (escribe SI): ").strip().upper()
    return ok == "SI"


def main() -> int:
    modo = _pedir_modo()
    archivo = pedir_ruta_excel()
    col_letter = pedir_columna()
    col_idx = column_index_from_string(col_letter)

    print(f"\nArchivo: {archivo}")
    print(f"Columna: {col_letter}")

    if modo == "1":
        print("Reanclando (XML interno; no se borran imágenes WMF ni media)…\n")
        try:
            total, copia = reanclar_columna_en_xlsx(archivo, col_letter)
        except OSError as e:
            print(f"Error al guardar (¿Excel abierto?): {e}", file=sys.stderr)
            return 1
        print(f"\nRespaldo: {copia}")
        print(f"Guardado: {archivo}")
        print(
            f"Fotos ajustadas: {total['ajustadas']} | "
            f"ya correctas: {total['ok']} | otra col: {total['otra_col']}"
        )
        return 0 if total["ajustadas"] or total["ok"] else 2

    if not _confirmar_modo_urls_si_hay_imagenes(archivo):
        print("Cancelado. Usa modo 1 para reanclar sin perder imágenes.")
        return 1

    print("Cargando con openpyxl (solo para insertar desde URL)…\n")
    aplicar_lectura_tolerante_imagenes()
    wb = load_workbook_seguro(archivo)

    total = _insertar_desde_urls(
        wb, col_letter, col_idx, os.path.dirname(archivo)
    )
    try:
        copia = guardar_workbook_con_respaldo(wb, archivo)
    except OSError as e:
        print(f"Error al guardar (¿Excel abierto?): {e}", file=sys.stderr)
        return 1
    print(f"\nRespaldo: {copia}")
    print(f"Guardado: {archivo}")
    print(f"Insertadas desde URL: {total['insertadas']}")
    return 0 if total["insertadas"] else 2


if __name__ == "__main__":
    sys.exit(main())
