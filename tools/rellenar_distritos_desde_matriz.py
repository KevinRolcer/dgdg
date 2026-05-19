"""
Rellena columnas C (Distrito Federal) y D (Distrito Local) según columna E (Municipio),
cruzando con docs/distrito_federal_y_local_por_municipio_2025.txt.
Solo la primera hoja; filas desde la 9. Rutas de entrada: extract_img.RUTA_CARPETA y NOMBRE_ARCHIVO.
"""
from __future__ import annotations

import os
import sys
import unicodedata
from pathlib import Path

import openpyxl

# Misma configuración que extract_img.py
from extract_img import NOMBRE_ARCHIVO, RUTA_CARPETA

ROOT = Path(__file__).resolve().parents[1]
MATRIZ_TXT = ROOT / "docs" / "distrito_federal_y_local_por_municipio_2025.txt"

# Nombre en extract_img; si no existe, se prueban variantes habituales del archivo.
_CANDIDATOS_EXCEL = [
    NOMBRE_ARCHIVO,
    "Aspirnates Presidentes Municipales.xlsx",
    "Perfiles Aspirnates Presidentes Municipales.xlsx",
    "Perfiles Aspirantes Presidentes Municipales.xlsx",
]


def _resolver_excel() -> str | None:
    for nombre in _CANDIDATOS_EXCEL:
        ruta = os.path.join(RUTA_CARPETA, nombre)
        if os.path.isfile(ruta):
            return ruta
    return None

FILA_INICIO = 9
COL_DTTO_FED = 3  # C
COL_DTTO_LOC = 4  # D
COL_MUNICIPIO = 5  # E


def normalizar_clave(texto: str) -> str:
    s = unicodedata.normalize("NFD", str(texto).strip().upper())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def cargar_matriz(path: Path) -> dict[str, tuple[str, str]]:
    if not path.is_file():
        raise FileNotFoundError(f"No existe la matriz TXT: {path}")
    out: dict[str, tuple[str, str]] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("-"):
            continue
        if line.startswith(("Lista ", "Columnas ", "Filas ", "Formato ", "Municipio,")):
            continue
        parts = line.split(",", 2)
        if len(parts) != 3:
            continue
        mun, dfed, dloc = parts[0], parts[1].strip(), parts[2].strip()
        out[normalizar_clave(mun)] = (dfed, dloc)
    return out


def main() -> int:
    archivo = _resolver_excel()
    if not archivo:
        print(
            "Error: no se encontró el Excel en "
            f"{RUTA_CARPETA!r} (probados: {_CANDIDATOS_EXCEL})",
            file=sys.stderr,
        )
        return 1
    matriz = cargar_matriz(MATRIZ_TXT)
    wb = openpyxl.load_workbook(archivo)
    ws = wb.worksheets[0]
    actualizados = 0
    sin_coincidencia: list[tuple[int, str]] = []
    vacios = 0

    for row in range(FILA_INICIO, ws.max_row + 1):
        raw = ws.cell(row=row, column=COL_MUNICIPIO).value
        if raw is None or (isinstance(raw, str) and not raw.strip()):
            vacios += 1
            continue
        clave = normalizar_clave(raw)
        if not clave:
            continue
        if clave not in matriz:
            sin_coincidencia.append((row, str(raw).strip()))
            continue
        dfed, dloc = matriz[clave]
        ws.cell(row=row, column=COL_DTTO_FED, value=dfed)
        ws.cell(row=row, column=COL_DTTO_LOC, value=dloc)
        actualizados += 1

    wb.save(archivo)
    print(f"Guardado: {archivo}")
    print(f"Filas con municipio actualizadas (C y D): {actualizados}")
    print(f"Celdas E vacías omitidas (fila {FILA_INICIO}+): {vacios}")
    if sin_coincidencia:
        print("Sin coincidencia en matriz (revisar nombre de municipio):")
        for r, m in sin_coincidencia[:40]:
            print(f"  fila {r}: {m!r}")
        if len(sin_coincidencia) > 40:
            print(f"  ... y {len(sin_coincidencia) - 40} más")
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
